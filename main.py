# Taeyong Lee
import openpyxl
from tabulate import tabulate
import random
import time

team_ratings = openpyxl.load_workbook("TeamRatings.xlsx")
sheet = team_ratings['Sheet1']


def start():
    team_id = [
        ["Atlanta Hawks", 1],
        ["Boston Celtics", 2],
        ["Brooklyn Nets", 3],
        ["Charlotte Hornets", 4],
        ["Chicago Bulls", 5],
        ["Cleveland Cavaliers", 6],
        ["Dallas Mavericks", 7],
        ["Denver Nuggets", 8],
        ["Detroit Pistons", 9],
        ["Golden State Warriors", 10],
        ["Houston Rockets", 11],
        ["Indiana Pacers", 12],
        ["Los Angeles Clippers", 13],
        ["Los Angeles Lakers", 14],
        ["Memphis Grizzlies", 15],
        ["Miami Heat", 16],
        ["Milwaukee Bucks", 17],
        ["Minnesota Timberwolves", 18],
        ["New Orleans Pelicans", 19],
        ["New York Knicks", 20],
        ["Oklahoma City Thunder", 21],
        ["Orlando Magic", 22],
        ["Philadelphia 76ers", 23],
        ["Phoenix Suns", 24],
        ["Portland Trail Blazers", 25],
        ["Sacramento Kings", 26],
        ["San Antonio Spurs", 27],
        ["Toronto Raptors", 28],
        ["Utah Jazz", 29],
        ["Washington Wizards", 30]
    ]
    head = ["Team Name", "ID"]

    print("Here is the list of teams and their id's to get you started: ")
    print(tabulate(team_id, headers=head, tablefmt="simple_grid"))
    print()
    user_choice = int(input(
        "Would like you to simulate the whole playoffs or one series? Enter 1 for playoffs, 2 for one series: "))
    while user_choice != 1 and user_choice != 2:
        print("Invalid input. Try again? ")
        user_choice = int(input(
            "Would you like to simulate the whole playoffs or one series? Enter 1 for playoffs, 2 for one series: "))

    if user_choice == 2:
        series = Series()
        series.simulateSeries()
    elif user_choice == 1:
        playoffs = Playoffs()
        playoffs.generateTeams()
        playoffs.simulatePlayoffs()


class Team:
    def __init__(self, id):
        self.id = id
        self.conference = None
        self.rating = None
        self.team_name = None
        self.seed = None

    def getAttributes(self):
        for i in sheet.iter_rows():
            id = i[0].value
            if id == self.id:
                self.team_name = i[1].value
                self.conference = i[2].value
                self.rating = i[15].value

    def printAttributes(self):
        print(f"Team Name: {self.team_name}")
        print(f"Team Conference: {self.conference}")
        print(f"Team Rating: {self.rating}")

    def __str__(self):
        return str(self.team_name)


class Series():
    def __init__(self):
        self.winner = None
        self.winner_id = None
        self.loser = None
        self.loser_id = None
        self.games = 0
        self.round_name = None
        self.conference = None

    def simulateSeries(self, team1_id=None, team2_id=None):
        if team1_id is None:
            team1_id = int(input("Enter in the id for the first team: "))
        team1 = Team(team1_id)
        team1.getAttributes()
        if team2_id is None:
            team2_id = int(input("Enter in the id for the second team: "))
        team2 = Team(team2_id)
        team2.getAttributes()

        team1_wins = 0
        team2_wins = 0
        team1_low = team1.rating - 5
        team1_high = team1.rating + 5
        team2_low = team2.rating - 5
        team2_high = team2.rating + 5

        while team1_wins < 4 and team2_wins < 4:
            team1_score = random.uniform(team1_low, team1_high)
            team2_score = random.uniform(team2_low, team2_high)
            self.games += 1
            if team1_score > team2_score:
                team1_wins += 1
                print(f"Game {self.games} winner: {team1}")

            else:
                team2_wins += 1
                print(f"Game {self.games} winner: {team2}")
            time.sleep(0.5)

        if team1_wins == 4:
            self.winner = team1
            self.winner_id = team1.id
            self.loser = team2
            self.loser_id = team2.id

        else:
            self.winner = team2
            self.winner_id = team2.id
            self.loser = team1
            self.loser_id = team1.id

        # needs fixing
        if self.conference is None:
            print(
                f"{self.winner} win the series against the {self.loser} in {self.games} games!")
        elif self.conference == "Finals":
            print(
                f"{self.winner} win the NBA FINALS against the {self.loser} in {self.games} games!")
        else:
            print(
                f"{self.winner} win the {self.conference} {self.round_name} against the {self.loser} in {self.games} games!")
            print()


class Playoffs:
    def __init__(self):
        self.teams_east = []
        self.teams_west = []

    def generateTeams(self):
        while len(self.teams_east) < 8:
            random_east_id = random.randint(1, 30)
            for i in sheet.iter_rows():
                id = i[0].value
                if id == random_east_id:
                    conference = i[2].value
                    if conference == "East" and id not in self.teams_east:
                        self.teams_east.append(id)
        while len(self.teams_west) < 8:
            random_west_id = random.randint(1, 30)
            for i in sheet.iter_rows():
                id = i[0].value
                if id == random_west_id:
                    conference = i[2].value
                    if conference == "West" and id not in self.teams_west:
                        self.teams_west.append(id)

    def simulatePlayoffs(self):
        # East
        current_round_east = self.teams_east
        current_round_winners_east = []

        rounds_finished_east = 0
        while rounds_finished_east < 3:
            i = 0
            j = len(current_round_east) - 1
            while i < j:
                series = Series()
                series.conference = "Eastern"
                if rounds_finished_east == 0:
                    series.round_name = "Conference Round One"
                elif rounds_finished_east == 1:
                    series.round_name = "Conference Semi Finals"
                elif rounds_finished_east == 2:
                    series.round_name = "Conference Finals"
                series.simulateSeries(
                    current_round_east[i], current_round_east[j])
                current_round_winners_east.append(series.winner_id)
                i += 1
                j -= 1
            rounds_finished_east += 1
            current_round_east = current_round_winners_east
            current_round_winners_east = []

        # West
        current_round_west = self.teams_west
        current_round_winners_west = []

        rounds_finished_west = 0
        while rounds_finished_west < 3:
            i = 0
            j = len(current_round_west) - 1
            while i < j:
                series = Series()
                series.conference = "Western"
                if rounds_finished_west == 0:
                    series.round_name = "Conference Round One"
                elif rounds_finished_west == 1:
                    series.round_name = "Conference Semi Finals"
                elif rounds_finished_west == 2:
                    series.round_name = "Conference Finals"
                series.simulateSeries(
                    current_round_west[i], current_round_west[j])
                current_round_winners_west.append(series.winner_id)
                i += 1
                j -= 1
            rounds_finished_west += 1
            current_round_west = current_round_winners_west
            current_round_winners_west = []

        # Finals
        finals = Series()
        finals.conference = "Finals"
        finals.simulateSeries(current_round_east[0], current_round_west[0])

    def show(self):
        print(self.teams_east)
        print(self.teams_west)


if __name__ == "__main__":
    start = start()
